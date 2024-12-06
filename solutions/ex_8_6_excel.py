from pathlib import Path
from datetime import date

from openpyxl import load_workbook

current_path = Path().resolve()
excel_file = current_path.parent / "docs" / "Financial Sample.xlsx"

wb = load_workbook(excel_file)
ws = wb.active

ws_q4 = wb.create_sheet("Q4 2014")
ws_q4.append([cell.value for cell in ws[1]])

date_start = date(2014, 10, 1)
date_end = date(2014, 12, 31)

for row in ws.iter_rows(min_row=2):
    profit = row[11].value
    dt = row[12].value.date()

    if profit > 100_000 and date_start < dt < date_end:
        ws_q4.append([cell.value for cell in row])

wb.save("modified.xlsx")
